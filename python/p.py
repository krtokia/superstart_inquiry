# -*- coding:utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection
from urllib.parse import unquote, quote, quote_plus, urlencode
from bs4 import BeautifulSoup
import collections
import re
import json
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding='utf-8')

wb = Workbook()
ws = wb.active


def ydef(data):
    data_length = len(data[list(data.keys())[0]])
    key_length = list(data.keys())
    row = 1
    col = 1
    scol = 1
    srow = 1
    snap = 1
    for i in list(data.keys()):
        yield {"row": row, "col": col, "value": str(i)}
        ws.cell(row=row, column=col).font = Font(color="000000", bold=True)
        row += 1
        for j in data[i]:
            scol = col
            if type(j) is str:
                yield {"row": row, "col": scol, "value": str(j)}
                scol += 1
            else:
                snap = row
                yield "SECOND SEPERATOR"
                chk = False
                for k in list(j.keys()):
                    yield "SEPERATOR"
                    yield {"row": 2, "col": scol, "value": k}
                    ws.cell(row=2, column=scol).font = Font(
                        color="000000", bold=True)
                    srow = row
                    for l in j[k]:
                        srow += 1
                        snap += 1
                        yield {"row": srow, "col": scol, "value": l}
                    scol += 1
                    if len(j[k]) > 0:
                        chk = True
                        snap = snap-1
                    # snap = snap - len(j[k])
                # if chk:
                #     snap = snap-1
                # row = snap-1
                # row = row-1
                row = snap
                yield "SECOND SEPERATOR /"
            row += 1
        col = scol+1
        # scol = col
        row = 1


with open("/data/new_json6.json", 'r', encoding='utf-8') as f:
    ori_data = json.load(f)


fin_data = {}
for d in ori_data:
    if d['column'] not in fin_data:
        fin_data[d['column']] = []
    fin_data[d['column']] = fin_data[d['column']] + d['data']
for i in ydef(fin_data):
    print(i)
    if "row" in i:
        ws.cell(row=i['row'], column=i['col'], value=str(i['value']))


wb.save("t1.xlsx")
